"""
Statistics and reporting module for travel invoice system.

Provides functionality for:
- Monthly invoice statistics
- Traveler statistics
- Invoice type statistics
- Excel report export
"""

import os
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from dataclasses import dataclass
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from loguru import logger


@dataclass
class MonthlyStats:
    """Monthly statistics data."""
    year: int
    month: int
    total_amount: float
    invoice_count: int
    by_type: Dict[str, float]
    by_traveler: Dict[str, float]


@dataclass
class TravelerStats:
    """Traveler statistics data."""
    name: str
    total_amount: float
    invoice_count: int
    trip_count: int
    by_type: Dict[str, float]
    by_month: Dict[str, float]


@dataclass
class TypeStats:
    """Invoice type statistics data."""
    type_name: str
    total_amount: float
    invoice_count: int
    avg_amount: float


class InvoiceStatistics:
    """
    Calculate and generate statistics for invoices.
    """

    def __init__(self, invoices_dir: str = "invoices"):
        """Initialize statistics calculator.

        Args:
            invoices_dir: Directory containing organized invoices
        """
        self.invoices_dir = Path(invoices_dir)

    def get_monthly_stats(self, months: int = 12) -> List[MonthlyStats]:
        """
        Get monthly statistics for the past N months.

        Args:
            months: Number of months to include

        Returns:
            List of monthly statistics
        """
        stats = []
        today = date.today()

        for i in range(months):
            # Calculate month boundaries
            month_date = today - timedelta(days=30 * i)
            year = month_date.year
            month = month_date.month

            # Gather data for this month
            total_amount = 0.0
            invoice_count = 0
            by_type = defaultdict(float)
            by_traveler = defaultdict(float)

            month_dir = self.invoices_dir / str(year) / f"{month:02d}"
            if not month_dir.exists():
                continue

            for type_dir in month_dir.iterdir():
                if not type_dir.is_dir():
                    continue

                type_name = type_dir.name
                for pdf_file in type_dir.glob("*.pdf"):
                    # Extract amount from filename
                    amount = self._extract_amount_from_filename(pdf_file.name)
                    if amount:
                        total_amount += amount
                        invoice_count += 1
                        by_type[type_name] += amount

                    # Extract traveler from filename
                    traveler = self._extract_traveler_from_filename(pdf_file.name)
                    if traveler:
                        by_traveler[traveler] += amount

            stats.append(MonthlyStats(
                year=year,
                month=month,
                total_amount=total_amount,
                invoice_count=invoice_count,
                by_type=dict(by_type),
                by_traveler=dict(by_traveler)
            ))

        return stats

    def get_traveler_stats(self) -> List[TravelerStats]:
        """
        Get statistics grouped by traveler.

        Returns:
            List of traveler statistics
        """
        traveler_data = defaultdict(lambda: {
            'total_amount': 0.0,
            'invoice_count': 0,
            'by_type': defaultdict(float),
            'by_month': defaultdict(float)
        })

        # Scan all invoice files
        for pdf_file in self.invoices_dir.rglob("*.pdf"):
            amount = self._extract_amount_from_filename(pdf_file.name)
            if not amount:
                continue

            traveler = self._extract_traveler_from_filename(pdf_file.name)
            if not traveler:
                traveler = "未知"

            # Get invoice type from directory path
            invoice_type = "其他"
            parts = pdf_file.relative_to(self.invoices_dir).parts
            if len(parts) >= 3:
                invoice_type = parts[2]

            # Get month from path
            month_key = ""
            if len(parts) >= 2:
                month_key = f"{parts[0]}-{parts[1]}"

            traveler_data[traveler]['total_amount'] += amount
            traveler_data[traveler]['invoice_count'] += 1
            traveler_data[traveler]['by_type'][invoice_type] += amount
            if month_key:
                traveler_data[traveler]['by_month'][month_key] += amount

        # Count trips from trips directory
        trip_counts = self._count_trips_by_traveler()

        # Convert to dataclass objects
        stats = []
        for name, data in traveler_data.items():
            stats.append(TravelerStats(
                name=name,
                total_amount=data['total_amount'],
                invoice_count=data['invoice_count'],
                trip_count=trip_counts.get(name, 0),
                by_type=dict(data['by_type']),
                by_month=dict(data['by_month'])
            ))

        # Sort by total amount descending
        stats.sort(key=lambda x: x.total_amount, reverse=True)
        return stats

    def get_type_stats(self) -> List[TypeStats]:
        """
        Get statistics grouped by invoice type.

        Returns:
            List of type statistics
        """
        type_data = defaultdict(lambda: {'total_amount': 0.0, 'invoice_count': 0})

        for pdf_file in self.invoices_dir.rglob("*.pdf"):
            amount = self._extract_amount_from_filename(pdf_file.name)
            if not amount:
                continue

            # Get invoice type from directory path
            invoice_type = "其他"
            parts = pdf_file.relative_to(self.invoices_dir).parts
            if len(parts) >= 3:
                invoice_type = parts[2]

            type_data[invoice_type]['total_amount'] += amount
            type_data[invoice_type]['invoice_count'] += 1

        # Convert to dataclass objects
        stats = []
        for type_name, data in type_data.items():
            avg_amount = data['total_amount'] / data['invoice_count'] if data['invoice_count'] > 0 else 0
            stats.append(TypeStats(
                type_name=type_name,
                total_amount=data['total_amount'],
                invoice_count=data['invoice_count'],
                avg_amount=avg_amount
            ))

        # Sort by total amount descending
        stats.sort(key=lambda x: x.total_amount, reverse=True)
        return stats

    def _extract_amount_from_filename(self, filename: str) -> Optional[float]:
        """Extract amount from filename."""
        import re
        # Look for pattern: 123.45 or 123,45
        patterns = [
            r'_(\d+\.\d{2})_',
            r'_(\d+,\d{2})_',
            r'_(\d+)\.\d{2}_',
        ]
        for pattern in patterns:
            match = re.search(pattern, filename)
            if match:
                try:
                    return float(match.group(1).replace(',', '.'))
                except ValueError:
                    continue
        return None

    def _extract_traveler_from_filename(self, filename: str) -> Optional[str]:
        """Extract traveler name from filename."""
        import re
        # Look for pattern: _王春晖.pdf or _王春晖_
        patterns = [
            r'_([\u4e00-\u9fa5]{2,4})\.pdf$',
            r'_([\u4e00-\u9fa5]{2,4})_',
        ]
        for pattern in patterns:
            match = re.search(pattern, filename)
            if match:
                return match.group(1)
        return None

    def _count_trips_by_traveler(self) -> Dict[str, int]:
        """Count trips by traveler from trips directory."""
        trip_counts = {}
        trips_dir = Path("trips")
        if not trips_dir.exists():
            return trip_counts

        for traveler_dir in trips_dir.iterdir():
            if not traveler_dir.is_dir() or traveler_dir.name.startswith('.'):
                continue

            count = sum(1 for d in traveler_dir.iterdir()
                       if d.is_dir() and d.name != '普通打车')
            trip_counts[traveler_dir.name] = count

        return trip_counts


class ExcelReportGenerator:
    """
    Generate Excel reports for invoice statistics.
    """

    def __init__(self):
        """Initialize Excel report generator."""
        self.stats = InvoiceStatistics()

    def create_monthly_report(self, output_path: str, months: int = 12):
        """
        Create monthly statistics Excel report.

        Args:
            output_path: Path to save the Excel file
            months: Number of months to include
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "月度统计"

        # Get statistics data
        monthly_stats = self.stats.get_monthly_stats(months)
        if not monthly_stats:
            logger.warning("No monthly statistics data available")
            return

        # Write header
        headers = ["月份", "总金额", "发票数量", "机票", "火车", "打车/接送机", "住宿", "餐饮", "其他"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            self._style_header(cell)

        # Write data
        for row_idx, stat in enumerate(monthly_stats, start=2):
            month_str = f"{stat.year}-{stat.month:02d}"
            ws.cell(row=row_idx, column=1, value=month_str)
            ws.cell(row=row_idx, column=2, value=stat.total_amount)
            ws.cell(row=row_idx, column=3, value=stat.invoice_count)

            # Type amounts
            type_cols = {
                "交通": 4,  # Combined 交通
                "住宿": 6,
                "餐饮": 7,
                "其他": 8
            }

            # 交通 combines 机票, 火车, 打车, 接送机
            transport_amount = (
                stat.by_type.get("机票", 0) +
                stat.by_type.get("火车", 0) +
                stat.by_type.get("打车", 0) +
                stat.by_type.get("接送机", 0)
            )
            ws.cell(row=row_idx, column=type_cols["交通"], value=transport_amount)
            ws.cell(row=row_idx, column=type_cols["住宿"], value=stat.by_type.get("住宿", 0))
            ws.cell(row=row_idx, column=type_cols["餐饮"], value=stat.by_type.get("餐饮", 0))
            ws.cell(row=row_idx, column=type_cols["其他"], value=stat.by_type.get("其他", 0))

        # Auto-fit columns
        self._auto_fit_columns(ws)

        # Add summary sheet
        self._add_summary_sheet(wb, monthly_stats)

        wb.save(output_path)
        logger.info(f"Monthly report saved to {output_path}")

    def create_traveler_report(self, output_path: str):
        """
        Create traveler statistics Excel report.

        Args:
            output_path: Path to save the Excel file
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "出差人统计"

        # Get statistics data
        traveler_stats = self.stats.get_traveler_stats()

        # Write header
        headers = ["出差人", "总金额", "发票数量", "行程数", "机票", "火车", "打车/接送机", "住宿", "餐饮"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            self._style_header(cell)

        # Write data
        for row_idx, stat in enumerate(traveler_stats, start=2):
            ws.cell(row=row_idx, column=1, value=stat.name)
            ws.cell(row=row_idx, column=2, value=stat.total_amount)
            ws.cell(row=row_idx, column=3, value=stat.invoice_count)
            ws.cell(row=row_idx, column=4, value=stat.trip_count)

            # Type amounts
            transport = (
                stat.by_type.get("机票", 0) +
                stat.by_type.get("火车", 0) +
                stat.by_type.get("打车", 0) +
                stat.by_type.get("接送机", 0)
            )
            ws.cell(row=row_idx, column=5, value=stat.by_type.get("机票", 0))
            ws.cell(row=row_idx, column=6, value=stat.by_type.get("火车", 0))
            ws.cell(row=row_idx, column=7, value=transport)
            ws.cell(row=row_idx, column=8, value=stat.by_type.get("住宿", 0))
            ws.cell(row=row_idx, column=9, value=stat.by_type.get("餐饮", 0))

        # Auto-fit columns
        self._auto_fit_columns(ws)

        wb.save(output_path)
        logger.info(f"Traveler report saved to {output_path}")

    def create_type_report(self, output_path: str):
        """
        Create invoice type statistics Excel report.

        Args:
            output_path: Path to save the Excel file
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "类型统计"

        # Get statistics data
        type_stats = self.stats.get_type_stats()

        # Write header
        headers = ["类型", "总金额", "发票数量", "平均金额"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            self._style_header(cell)

        # Write data
        for row_idx, stat in enumerate(type_stats, start=2):
            ws.cell(row=row_idx, column=1, value=stat.type_name)
            ws.cell(row=row_idx, column=2, value=stat.total_amount)
            ws.cell(row=row_idx, column=3, value=stat.invoice_count)
            ws.cell(row=row_idx, column=4, value=stat.avg_amount)

        # Auto-fit columns
        self._auto_fit_columns(ws)

        wb.save(output_path)
        logger.info(f"Type report saved to {output_path}")

    def create_comprehensive_report(self, output_path: str, months: int = 12):
        """
        Create comprehensive Excel report with multiple sheets.

        Args:
            output_path: Path to save the Excel file
            months: Number of months for monthly stats
        """
        wb = openpyxl.Workbook()

        # Create different sheets
        self._create_monthly_sheet(wb, months)
        self._create_traveler_sheet(wb)
        self._create_type_sheet(wb)

        wb.save(output_path)
        logger.info(f"Comprehensive report saved to {output_path}")

    def _create_monthly_sheet(self, wb, months):
        """Create monthly statistics sheet."""
        ws = wb.create_sheet("月度统计")

        monthly_stats = self.stats.get_monthly_stats(months)
        if not monthly_stats:
            return

        # Header
        headers = ["月份", "总金额", "发票数量", "交通", "住宿", "餐饮"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            self._style_header(cell)

        # Data
        for row_idx, stat in enumerate(reversed(monthly_stats), start=2):
            month_str = f"{stat.year}-{stat.month:02d}"
            transport = (
                stat.by_type.get("机票", 0) + stat.by_type.get("火车", 0) +
                stat.by_type.get("打车", 0) + stat.by_type.get("接送机", 0)
            )

            ws.cell(row=row_idx, column=1, value=month_str)
            ws.cell(row=row_idx, column=2, value=stat.total_amount)
            ws.cell(row=row_idx, column=3, value=stat.invoice_count)
            ws.cell(row=row_idx, column=4, value=transport)
            ws.cell(row=row_idx, column=5, value=stat.by_type.get("住宿", 0))
            ws.cell(row=row_idx, column=6, value=stat.by_type.get("餐饮", 0))

        self._auto_fit_columns(ws)

    def _create_traveler_sheet(self, wb):
        """Create traveler statistics sheet."""
        ws = wb.create_sheet("出差人统计")

        traveler_stats = self.stats.get_traveler_stats()

        # Header
        headers = ["出差人", "总金额", "发票数", "行程数"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            self._style_header(cell)

        # Data
        for row_idx, stat in enumerate(traveler_stats, start=2):
            ws.cell(row=row_idx, column=1, value=stat.name)
            ws.cell(row=row_idx, column=2, value=stat.total_amount)
            ws.cell(row=row_idx, column=3, value=stat.invoice_count)
            ws.cell(row=row_idx, column=4, value=stat.trip_count)

        self._auto_fit_columns(ws)

    def _create_type_sheet(self, wb):
        """Create invoice type statistics sheet."""
        ws = wb.create_sheet("类型统计")

        type_stats = self.stats.get_type_stats()

        # Header
        headers = ["类型", "总金额", "发票数", "平均金额"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            self._style_header(cell)

        # Data
        for row_idx, stat in enumerate(type_stats, start=2):
            ws.cell(row=row_idx, column=1, value=stat.type_name)
            ws.cell(row=row_idx, column=2, value=stat.total_amount)
            ws.cell(row=row_idx, column=3, value=stat.invoice_count)
            ws.cell(row=row_idx, column=4, value=stat.avg_amount)

        self._auto_fit_columns(ws)

    def _add_summary_sheet(self, wb, monthly_stats):
        """Add summary sheet with totals."""
        ws = wb.create_sheet("汇总")

        total_amount = sum(s.total_amount for s in monthly_stats)
        total_invoices = sum(s.invoice_count for s in monthly_stats)

        # Title
        ws.cell(row=1, column=1, value="统计汇总")
        ws.cell(row=1, column=1).font = Font(size=16, bold=True)

        # Summary data
        ws.cell(row=3, column=1, value="统计周期:")
        ws.cell(row=3, column=2, value=f"{monthly_stats[-1].year}-{monthly_stats[-1].month:02d} 至 {monthly_stats[0].year}-{monthly_stats[0].month:02d}")
        ws.cell(row=4, column=1, value="总金额:")
        ws.cell(row=4, column=2, value=total_amount)
        ws.cell(row=4, column=2).number_format = '#,##0.00'
        ws.cell(row=5, column=1, value="发票总数:")
        ws.cell(row=5, column=2, value=total_invoices)

    def _style_header(self, cell):
        """Style header cell."""
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def _auto_fit_columns(self, ws):
        """Auto-fit column widths."""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
